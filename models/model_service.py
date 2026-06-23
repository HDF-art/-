import os
import torch
import torch.nn.functional as F
from torchvision import transforms
from PIL import Image
import numpy as np
from flask import Flask, request, jsonify
import logging
import sys
import json
import base64

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# 导入SepResNet模型
from SepResNet import SepResNet

# 配置日志
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    handlers=[logging.StreamHandler()])
logger = logging.getLogger('model_service')

# 初始化Flask应用
app = Flask(__name__)

# 配置参数
MODEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'global_model.pth')
PORT = 5000
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
logger.info(f'Using device: {device}')

# 图像预处理转换
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(
        mean=[0.485, 0.456, 0.406],
        std=[0.229, 0.224, 0.225]
    )
])

# 病虫害类别(必须与训练时的类别数量和顺序一致)
# 根据模型实际训练的15个类别
CLASSES = [
    '稻象甲科',      # 0 - Curculionidae
    '飞虱科',        # 1 - Delphacidae
    '叶蝉科',        # 2 - Cicadellidae
    '管蓟马科',      # 3 - Phlaeothripidae
    '瘿蚊科',        # 4 - Cecidomyiidae
    '弄蝶科',        # 5 - Hesperiidae
    '草螟科',        # 6 - Crambidae
    '秆蝇科',        # 7 - Chloropidae
    '水蝇科',        # 8 - Ephydridae
    '夜蛾科',        # 9 - Noctuidae
    '蓟马科',        # 10 - Thripidae
    '细菌性白叶枯病', # 11 - Bacterialblight
    '稻瘟病',        # 12 - Blast
    '褐斑病',        # 13 - Brownspot
    '东格鲁病'       # 14 - Tungro
]

# 加载模型
model = None

def load_model():
    global model
    try:
        logger.info(f'Loading model from {MODEL_PATH}')
        
        # 从checkpoint推断模型参数
        # 根据错误信息,checkpoint的参数是:
        # - Conv_r = 8 (从lora_E形状推断)
        # - Linear_r = 12 (从fc.lora_E形状推断)
        # - num_classes = 15 (从fc.weight形状推断)
        
        # 初始化SepResNet模型 - 使用与训练时相同的参数
        model = SepResNet(num_classes=len(CLASSES), Conv_r=8, Linear_r=12)
        
        # 加载模型权重
        if os.path.exists(MODEL_PATH):
            checkpoint = torch.load(MODEL_PATH, map_location=device, weights_only=False)
            
            # 处理不同的权重保存格式
            if isinstance(checkpoint, dict):
                # 如果是字典格式,可能包含'model_state_dict'或'state_dict'等键
                if 'model_state_dict' in checkpoint:
                    state_dict = checkpoint['model_state_dict']
                    logger.info("使用键: 'model_state_dict'")
                elif 'state_dict' in checkpoint:
                    state_dict = checkpoint['state_dict']
                    logger.info("使用键: 'state_dict'")
                else:
                    # 直接作为state_dict使用
                    state_dict = checkpoint
                    logger.info("直接使用checkpoint作为state_dict")
            else:
                state_dict = checkpoint
                logger.info(f"Checkpoint类型: {type(checkpoint)}")
            
            # 加载权重 - 现在参数应该完全匹配
            missing_keys, unexpected_keys = model.load_state_dict(state_dict, strict=True)
            
            if missing_keys:
                logger.warning(f'Missing keys when loading model ({len(missing_keys)}): {missing_keys[:5]}...')
            if unexpected_keys:
                logger.warning(f'Unexpected keys when loading model ({len(unexpected_keys)}): {unexpected_keys[:5]}...')
            
            logger.info('Model weights loaded successfully')
        else:
            logger.warning(f'Model file not found at {MODEL_PATH}, using initialized weights')
        
        # 移动到设备并设置为评估模式
        model.to(device)
        model.eval()
        logger.info('Model ready for inference')
        logger.info(f'Model configuration: num_classes={len(CLASSES)}, Conv_r=8, Linear_r=12')
        return True
    except Exception as e:
        logger.error(f'Error loading model: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return False

# 图像识别函数
def predict(image_path):
    try:
        # 加载和预处理图像
        image = Image.open(image_path).convert('RGB')
        image = transform(image).unsqueeze(0).to(device)
        
        # 模型推理 - SepResNet的forward需要mode参数
        with torch.no_grad():
            # 推理时使用mode='lora'或'all'来激活LoRA层
            # 'all'表示使用完整模型(原始权重+LoRA权重)
            outputs = model(image, mode='all')
            probabilities = F.softmax(outputs, dim=1)
            max_prob, predicted_class = torch.max(probabilities, 1)
            
            # 获取所有类别的概率
            all_probs = probabilities.squeeze().cpu().numpy()
            
        # 构建结果
        result = {
            'diseaseName': CLASSES[predicted_class.item()],
            'confidence': max_prob.item(),
            'details': [
                {'diseaseName': CLASSES[i], 'confidence': float(prob)}
                for i, prob in enumerate(all_probs)
            ]
        }
        
        logger.info(f'Prediction: {CLASSES[predicted_class.item()]} (confidence: {max_prob.item():.4f})')
        return result
    except Exception as e:
        logger.error(f'Error during prediction: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        raise

# REST API端点
@app.route('/predict', methods=['POST'])
def predict_endpoint():
    temp_path = None
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            logger.error('No file provided in request')
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        # 检查文件名
        if file.filename == '':
            logger.error('No file selected')
            return jsonify({'error': 'No file selected'}), 400
        
        # 使用更安全的临时文件路径
        import tempfile
        import uuid
        temp_dir = tempfile.gettempdir()
        temp_filename = f'model_predict_{uuid.uuid4().hex}.jpg'
        temp_path = os.path.join(temp_dir, temp_filename)
        
        # 保存上传的文件 - 使用stream方式确保完整保存
        logger.info(f'Saving uploaded file to: {temp_path}')
        try:
            # 读取文件内容
            file_content = file.read()
            logger.info(f'Read {len(file_content)} bytes from uploaded file')
            
            # 写入临时文件
            with open(temp_path, 'wb') as f:
                f.write(file_content)
            
            # 验证文件已保存且可读
            if not os.path.exists(temp_path):
                logger.error(f'File not saved: {temp_path}')
                return jsonify({'error': 'Failed to save file'}), 500
            
            file_size = os.path.getsize(temp_path)
            logger.info(f'File saved successfully, size: {file_size} bytes')
            
            if file_size == 0:
                logger.error('Saved file is empty (0 bytes)')
                return jsonify({'error': 'Uploaded file is empty'}), 400
                
        except Exception as e:
            logger.error(f'Error saving file: {str(e)}')
            import traceback
            logger.error(traceback.format_exc())
            return jsonify({'error': f'Failed to save file: {str(e)}'}), 500
        
        # 进行预测
        result = predict(temp_path)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f'Error in predict endpoint: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
    finally:
        # 清理临时文件
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                logger.info(f'Cleaned up temp file: {temp_path}')
            except Exception as e:
                logger.warning(f'Failed to clean up temp file: {str(e)}')

# 健康检查端点
@app.route('/health', methods=['GET'])
def health_check():
    if model is not None:
        return jsonify({'status': 'healthy', 'message': 'Model service is running'}), 200
    else:
        return jsonify({'status': 'unhealthy', 'message': 'Model not loaded'}), 503

# 直接预测函数，接受base64编码的图像数据
def direct_predict_from_base64(base64_image_data):
    """
    直接从base64编码的图像数据进行预测，不通过HTTP API
    """
    import tempfile
    import uuid
    
    # 解码base64图像数据
    image_bytes = base64.b64decode(base64_image_data)
    
    # 创建临时文件
    temp_dir = tempfile.gettempdir()
    temp_filename = f'direct_predict_{uuid.uuid4().hex}.jpg'
    temp_path = os.path.join(temp_dir, temp_filename)
    
    try:
        # 保存图像到临时文件
        with open(temp_path, 'wb') as f:
            f.write(image_bytes)
        
        # 调用预测函数
        result = predict(temp_path)
        return result
    finally:
        # 清理临时文件
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

# 命令行直接调用入口
def main_direct_call():
    """
    用于直接调用模型进行预测的入口函数
    从标准输入读取base64编码的图像数据
    """
    import sys
    
    # 从命令行参数或标准输入获取base64编码的图像数据
    if len(sys.argv) > 1:
        base64_image_data = sys.argv[1]
    else:
        # 从标准输入读取
        base64_image_data = sys.stdin.read().strip()
    
    if not base64_image_data:
        print(json.dumps({"error": "No image data provided"}))
        return
    
    try:
        # 执行预测
        result = direct_predict_from_base64(base64_image_data)
        # 输出JSON结果到标准输出
        print(json.dumps(result))
    except Exception as e:
        print(json.dumps({"error": str(e)}))

# 直接预测函数，接受图像文件路径

def main_predict_from_file():
    """
    用于直接对指定路径的图像文件进行预测的入口函数
    """
    import sys
    import traceback
    
    try:
        if len(sys.argv) < 3:
            print(json.dumps({"error": "No image file path provided"}))
            return
        
        image_path = sys.argv[2]
        
        if not os.path.exists(image_path):
            print(json.dumps({"error": f"Image file not found: {image_path}"}))
            return
        
        # 检查依赖项
        try:
            import torch
        except ImportError as e:
            print(json.dumps({"error": f"缺少依赖项: {str(e)}"}))
            return
        
        try:
            import torchvision
        except ImportError as e:
            print(json.dumps({"error": f"缺少依赖项: {str(e)}"}))
            return
        
        try:
            from PIL import Image
        except ImportError as e:
            print(json.dumps({"error": f"缺少依赖项: {str(e)}"}))
            return
        
        # 检查模型权重文件
        if not os.path.exists(MODEL_PATH):
            print(json.dumps({"error": f"模型权重文件不存在: {MODEL_PATH}"}))
            return
        
        # 执行预测
        result = predict(image_path)
        # 只输出JSON结果到标准输出，不输出任何其他信息
        print(json.dumps(result))
    except Exception as e:
        # 只输出JSON格式的错误信息，不输出任何其他信息
        print(json.dumps({"error": str(e)}))


# ==================== 草莓成熟度检测模型 ====================
STRAWBERRY_MODEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'YOLOv8-SP', 'runs', 'detect', 'runs', 'detect', 'strawberry2', 'weights', 'best.pt')
STRAWBERRY_CLASSES = ['unripe(未成熟)', 'ripe(成熟)', 'overripe(过熟)']
strawberry_model = None


def load_strawberry_model():
    """加载草莓成熟度检测模型（基于YOLOv8）"""
    global strawberry_model
    try:
        from ultralytics import YOLO
        logger.info(f'Loading strawberry model from {STRAWBERRY_MODEL_PATH}')
        strawberry_model = YOLO(STRAWBERRY_MODEL_PATH)
        logger.info('Strawberry model loaded successfully')
        return True
    except Exception as e:
        logger.error(f'Error loading strawberry model: {str(e)}')
        return False


def predict_strawberry(image_path):
    """草莓成熟度检测推理"""
    try:
        results = strawberry_model.predict(source=image_path, conf=0.1, verbose=False)
        detections = []
        for r in results:
            boxes = r.boxes
            for box in boxes:
                cls_id = int(box.cls[0])
                conf = float(box.conf[0])
                x1, y1, x2, y2 = box.xyxy[0].tolist()
                cls_name = STRAWBERRY_CLASSES[cls_id] if cls_id < len(STRAWBERRY_CLASSES) else f'class_{cls_id}'
                detections.append({
                    'className': cls_name,
                    'confidence': conf,
                    'bbox': [round(x1, 1), round(y1, 1), round(x2, 1), round(y2, 1)]
                })

        if not detections:
            return {
                'diseaseName': '未检测到草莓',
                'confidence': 0.0,
                'details': [],
                'detections': []
            }

        # 取置信度最高的检测结果作为主要识别
        best = max(detections, key=lambda x: x['confidence'])
        # 按类别聚合统计
        class_counts = {}
        for d in detections:
            class_counts[d['className']] = class_counts.get(d['className'], 0) + 1

        details = [
            {'diseaseName': name, 'confidence': count / len(detections)}
            for name, count in class_counts.items()
        ]

        return {
            'diseaseName': best['className'],
            'confidence': best['confidence'],
            'details': details,
            'detections': detections
        }
    except Exception as e:
        logger.error(f'Error during strawberry prediction: {str(e)}')
        raise


@app.route('/predict/strawberry', methods=['POST'])
def predict_strawberry_endpoint():
    """草莓成熟度检测端点"""
    temp_path = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        import tempfile, uuid
        temp_path = os.path.join(tempfile.gettempdir(),
                                  f'strawberry_{uuid.uuid4().hex}.jpg')
        file.save(temp_path)

        if strawberry_model is None:
            if not load_strawberry_model():
                return jsonify({'error': '草莓成熟度模型加载失败'}), 503

        result = predict_strawberry(temp_path)
        logger.info(f'Strawberry prediction: {result["diseaseName"]} ({result["confidence"]:.4f})')
        return jsonify(result)
    except Exception as e:
        logger.error(f'Error in strawberry predict endpoint: {str(e)}')
        return jsonify({'error': str(e)}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


# ==================== 环境预测模型（TCN-FECAM-iTransformer） ====================
ENV_MODEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'TCN-FECAM-iTransformer')
ENV_MODEL_PATH = os.path.join(ENV_MODEL_DIR, 'checkpoints', 'best_model.pth')
env_model = None
env_config = None
env_scaler = None


def load_env_prediction_model():
    """加载TCN-FECAM-iTransformer环境预测模型"""
    global env_model, env_config, env_scaler
    try:
        import sys as _sys
        import os as _os
        if ENV_MODEL_DIR not in _sys.path:
            _sys.path.insert(0, ENV_MODEL_DIR)
        # 切换工作目录到模型目录，使 Config 中的相对路径生效
        original_cwd = _os.getcwd()
        _os.chdir(ENV_MODEL_DIR)
        try:
            from config import Config
            from models import TCN_FECAM_iTransformer
            env_config = Config()
            # 将相对路径转为绝对路径
            if not _os.path.isabs(env_config.MODEL_SAVE_PATH):
                env_config.MODEL_SAVE_PATH = _os.path.join(ENV_MODEL_DIR, env_config.MODEL_SAVE_PATH)
            if not _os.path.isabs(env_config.RESULT_SAVE_DIR):
                env_config.RESULT_SAVE_DIR = _os.path.join(ENV_MODEL_DIR, env_config.RESULT_SAVE_DIR)
            checkpoint = torch.load(env_config.MODEL_SAVE_PATH,
                                    map_location=env_config.DEVICE,
                                    weights_only=False)
            env_model = TCN_FECAM_iTransformer(env_config).to(env_config.DEVICE)
            env_model.load_state_dict(checkpoint['model_state_dict'])
            env_model.eval()
            logger.info(f'Env prediction model loaded (Epoch {checkpoint.get("epoch", "?")}, '
                        f'Val Loss: {checkpoint.get("val_loss", "?"):.6f})')
            return True
        finally:
            _os.chdir(original_cwd)
    except Exception as e:
        logger.error(f'Error loading env prediction model: {str(e)}')
        return False


@app.route('/predict/env', methods=['POST'])
def predict_env_endpoint():
    """环境预测端点 - 接收历史数据JSON，返回未来预测"""
    try:
        data = request.get_json()
        if not data or 'history' not in data:
            return jsonify({'error': '缺少历史数据（history字段）'}), 400

        history = data['history']  # List[List[float]]，长度应为 SEQ_LEN
        if env_model is None:
            if not load_env_prediction_model():
                return jsonify({'error': '环境预测模型加载失败'}), 503

        import numpy as _np
        seq_len = env_config.SEQ_LEN
        n_features = len(env_config.FEATURE_COLUMNS)

        if len(history) < seq_len:
            return jsonify({'error': f'历史数据不足，需要至少 {seq_len} 个时间步'}), 400

        # 取最近 seq_len 个时间步
        recent = history[-seq_len:]
        x = _np.array(recent, dtype=_np.float32)
        if x.ndim != 2 or x.shape[1] != n_features:
            return jsonify({'error': f'输入特征维度应为 {n_features}'}), 400

        x_tensor = torch.from_numpy(x).unsqueeze(0).to(env_config.DEVICE)
        with torch.no_grad():
            pred = env_model(x_tensor).cpu().numpy()

        # pred shape: (1, pred_len, n_targets)
        target_names = env_config.TARGET_COLUMNS
        predictions = []
        for t in range(pred.shape[1]):
            step = {}
            for i, name in enumerate(target_names):
                step[name] = float(pred[0, t, i])
            predictions.append(step)

        result = {
            'predictions': predictions,
            'predLen': pred.shape[1],
            'targetNames': target_names,
            'message': f'已预测未来 {pred.shape[1]} 个时间步'
        }
        logger.info(f'Env prediction completed: {pred.shape[1]} steps')
        return jsonify(result)
    except Exception as e:
        logger.error(f'Error in env predict endpoint: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/env/info', methods=['GET'])
def env_info_endpoint():
    """获取环境预测模型信息"""
    if env_config is None:
        return jsonify({'error': '环境预测模型未加载'}), 503
    return jsonify({
        'seqLen': env_config.SEQ_LEN,
        'predLen': env_config.PRED_LEN,
        'featureColumns': env_config.FEATURE_COLUMNS,
        'targetColumns': env_config.TARGET_COLUMNS,
        'resampleMinutes': env_config.RESAMPLE_MINUTES
    })


@app.route('/env/parse-excel', methods=['POST'])
def parse_env_excel_endpoint():
    """解析Excel文件，提取时间序列数据用于环境预测"""
    temp_path = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        import tempfile, uuid
        temp_path = os.path.join(tempfile.gettempdir(),
                                f'env_excel_{uuid.uuid4().hex}.xlsx')
        file.save(temp_path)

        import openpyxl
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)

        # 合并所有sheet中的特征列
        feature_set = set(env_config.FEATURE_COLUMNS)
        time_col = None
        col_map = {}  # feature_name -> col_index

        # 遍历所有sheet，建立列名到(sheet, col_index)的映射
        sheet_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
            if not rows:
                continue
            header = rows[0]
            for idx, h in enumerate(header):
                if h is None:
                    continue
                h_str = str(h).strip()
                if h_str == '时间':
                    time_col = (sheet_name, idx)
                if h_str in feature_set:
                    col_map[h_str] = (sheet_name, idx)
            if len(rows) > 1:
                sheet_data[sheet_name] = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not col_map:
            return jsonify({'error': 'Excel中未找到匹配的特征列，需要包含：' + '、'.join(env_config.FEATURE_COLUMNS)}), 400

        # 检查是否所有特征列都找到了
        missing = feature_set - set(col_map.keys())
        if missing:
            return jsonify({'error': f'缺少特征列：{", ".join(missing)}'}), 400

        # 按时间合并数据
        # 先收集时间列
        time_data = {}
        if time_col:
            sheet_name, col_idx = time_col
            for row in sheet_data.get(sheet_name, []):
                if col_idx < len(row) and row[col_idx] is not None:
                    t = str(row[col_idx])
                    time_data[t] = {}

        # 填充各特征列
        for feat, (sheet_name, col_idx) in col_map.items():
            for row in sheet_data.get(sheet_name, []):
                if col_idx < len(row) and row[col_idx] is not None:
                    # 找最近的时间
                    time_idx = 0 if '时间' not in [h for h in row] else None
                    # 简单方式：按行索引匹配
                    pass

        # 简化处理：假设所有sheet按行对齐，直接按行索引合并
        # 获取最大行数
        max_rows = 0
        for sheet_name, rows in sheet_data.items():
            max_rows = max(max_rows, len(rows))

        history = []
        for i in range(max_rows):
            row_data = []
            for feat in env_config.FEATURE_COLUMNS:
                if feat in col_map:
                    sheet_name, col_idx = col_map[feat]
                    rows = sheet_data.get(sheet_name, [])
                    val = rows[i][col_idx] if i < len(rows) and col_idx < len(rows[i]) else None
                    if val is not None:
                        try:
                            row_data.append(float(val))
                        except (ValueError, TypeError):
                            row_data.append(0.0)
                    else:
                        row_data.append(0.0)
                else:
                    row_data.append(0.0)
            history.append(row_data)

        logger.info(f'Excel parsed: {len(history)} rows, {len(env_config.FEATURE_COLUMNS)} features')
        return jsonify({
            'history': history,
            'rows': len(history),
            'features': env_config.FEATURE_COLUMNS,
            'message': f'成功解析 {len(history)} 行数据'
        })
    except Exception as e:
        logger.error(f'Error parsing Excel: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


# 启动服务
if __name__ == '__main__':
    if load_model():
        # 预加载草莓成熟度检测模型
        try:
            load_strawberry_model()
        except Exception as e:
            logger.warning(f'草莓成熟度模型预加载失败（不影响主服务）: {str(e)}')
        # 预加载环境预测模型
        try:
            load_env_prediction_model()
        except Exception as e:
            logger.warning(f'环境预测模型预加载失败（不影响主服务）: {str(e)}')

        # 检查是否有命令行参数，如果有则执行直接调用模式
        import sys
        if len(sys.argv) > 1:
            if sys.argv[1] == '--direct':
                main_direct_call()
            elif sys.argv[1] == '--predict':
                main_predict_from_file()
        else:
            logger.info(f'Starting model service on port {PORT}')
            app.run(host='0.0.0.0', port=PORT, debug=False)
    else:
        logger.error('Failed to start model service: Model loading failed')
